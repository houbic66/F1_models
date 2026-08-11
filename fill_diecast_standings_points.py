from __future__ import annotations

import csv
import re
import time
from copy import copy
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

from audit_f1_wiki import (
    END_YEAR,
    START_YEAR,
    clean_wiki_text,
    driver_key,
    driver_similarity,
    expand_table,
    fetch_doc,
    norm,
    ratio,
    text_of,
)


INPUT = Path("outputs") / "wiki_audit" / "Diecast 2026 - doplneno z auditu - opraveno.xlsx"
OUTPUT = Path("outputs") / "wiki_audit" / "Diecast 2026 - doplneno z auditu - body 1976-2025.xlsx"
SOURCE_AUDIT = Path("outputs") / "wiki_audit" / "standings_fill_audit.csv"

BASE_URL = "https://en.wikipedia.org"
SHEET = "Overview"

COL_YEAR = 1
COL_TEAM = 2
COL_CAR = 3
COL_TYPE = 4
COL_DRIVER = 6
COL_D = 9
COL_DP = 10
COL_T = 11
COL_TP = 12
COL_PC = 13
MAX_COL = 16


@dataclass
class Standing:
    name: str
    pos: int | str | None
    points: int | float | None


@dataclass
class Entry:
    driver: str
    driver_key: str
    entrant: str
    entrant_url: str
    constructor: str
    constructor_url: str
    chassis: str
    chassis_url: str


def clean(value) -> str:
    if value is None:
        return ""
    value = str(value).replace("\xa0", " ")
    return re.sub(r"\s+", " ", value).strip()


def is_url(value) -> bool:
    return isinstance(value, str) and value.lower().startswith(("http://", "https://"))


def pc_is_zero(value) -> bool:
    return clean(value) in {"0", "0.0"}


def parse_points(text: str) -> int | float | None:
    text = clean_wiki_text(text)
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return None
    raw = match.group(0)
    value = float(raw) if "." in raw else int(raw)
    return int(value) if isinstance(value, float) and value.is_integer() else value


def parse_pos(text: str) -> int | str | None:
    text = clean_wiki_text(text)
    if not text:
        return None
    match = re.search(r"\d+", text)
    if match:
        return int(match.group(0))
    upper = text.upper()
    if "NC" in upper:
        return "NC"
    return text


def table_headers(matrix, header_row):
    return [clean_wiki_text(text_of(cell)) if cell is not None else "" for cell in matrix[header_row]]


def find_col(headers, *needles):
    lowered = [h.lower() for h in headers]
    for idx, header in enumerate(lowered):
        if all(needle in header for needle in needles):
            return idx
    return None


def find_any_col(headers, labels):
    lowered = [h.lower() for h in headers]
    for idx, header in enumerate(lowered):
        if any(label in header for label in labels):
            return idx
    return None


def find_points_col(headers):
    lowered = [h.lower().strip() for h in headers]
    for idx in range(len(lowered) - 1, -1, -1):
        if "points" in lowered[idx] or lowered[idx] == "pts":
            return idx
    return None


def link_from_cell(cell):
    if cell is None:
        return ""
    links = []
    for a in cell.xpath(".//a[@href]"):
        label = clean_wiki_text(text_of(a))
        href = a.get("href") or ""
        if href.startswith("/wiki/"):
            url = BASE_URL + href
            page = href.split("/wiki/", 1)[1]
        elif href.startswith(BASE_URL + "/wiki/"):
            url = href
            page = href.split("/wiki/", 1)[1]
        else:
            continue
        if ":" in page:
            continue
        if not label or label.startswith("["):
            continue
        links.append((label, url))
    if not links:
        return ""
    bad_sponsors = {
        "bwt",
        "oracle",
        "moneygram",
        "petronas",
        "aramco",
        "cognizant",
        "stake",
        "kick",
        "visa",
        "cash app",
        "orlen",
        "mild seven",
        "west",
    }
    for label, url in links:
        if norm(label) not in {norm(item) for item in bad_sponsors}:
            return url
    return ""


def parse_standings_table(doc, kind: str) -> dict[str, Standing]:
    entity_labels = ["driver"] if kind == "driver" else ["constructor", "manufacturer"]
    candidates = []
    for table in doc.xpath('//table[contains(@class,"wikitable")]'):
        matrix = expand_table(table)
        for header_row, row in enumerate(matrix[:4]):
            row_text = " ".join(clean_wiki_text(text_of(cell)) for cell in row if cell is not None)
            row_text_l = row_text.lower()
            if not any(label in row_text_l for label in entity_labels) or ("points" not in row_text_l and "pts" not in row_text_l):
                continue
            headers = table_headers(matrix, header_row)
            pos_col = find_col(headers, "pos")
            entity_col = find_any_col(headers, entity_labels)
            points_col = find_points_col(headers)
            if pos_col is None or entity_col is None or points_col is None:
                continue
            score = len(matrix) + points_col
            candidates.append((score, table, matrix, header_row, pos_col, entity_col, points_col))
            break
    if not candidates:
        return {}
    _, _, matrix, header_row, pos_col, entity_col, points_col = sorted(candidates, key=lambda item: item[0], reverse=True)[0]

    out = {}
    for row in matrix[header_row + 1 :]:
        if entity_col >= len(row) or row[entity_col] is None:
            continue
        name = clean_wiki_text(text_of(row[entity_col]))
        if not name or name.lower() in {"driver", "constructor"}:
            continue
        pos = parse_pos(text_of(row[pos_col])) if pos_col < len(row) and row[pos_col] is not None else None
        points = parse_points(text_of(row[points_col])) if points_col < len(row) and row[points_col] is not None else None
        key = driver_key(name) if kind == "driver" else norm(name)
        out.setdefault(key, Standing(name=name, pos=pos, points=points))
    return out


def best_driver_standing(row_driver: str, standings: dict[str, Standing]) -> tuple[Standing | None, float]:
    key = driver_key(row_driver)
    if key in standings:
        return standings[key], 1.0
    best = None
    best_score = 0.0
    for standing in standings.values():
        score = driver_similarity(row_driver, standing.name)
        if score > best_score:
            best = standing
            best_score = score
    return best, best_score


def parse_entries_with_links(year: int, doc) -> list[Entry]:
    from audit_f1_wiki import find_entries_table, driver_names_from_cell

    table = find_entries_table(doc)
    if table is None:
        return []
    matrix = expand_table(table)
    header_start = 0
    for i, row in enumerate(matrix):
        row_text = " ".join(clean_wiki_text(text_of(c)) for c in row if c is not None)
        if "Driver" in row_text and ("Entrant" in row_text or "Constructor" in row_text):
            header_start = i
            break
    matrix = matrix[header_start:]
    header_rows = 2 if any("Race drivers" in clean_wiki_text(text_of(c)) for c in matrix[0] if c is not None) else 1
    headers = []
    width = max((len(r) for r in matrix[:header_rows]), default=0)
    for col in range(width):
        parts = []
        for row in matrix[:header_rows]:
            cell = row[col] if col < len(row) else None
            if cell is not None:
                parts.append(clean_wiki_text(text_of(cell)))
        headers.append(" / ".join(part for part in parts if part))
    lowered = [h.lower() for h in headers]
    entrant_col = next((i for i, h in enumerate(lowered) if "entrant" in h), None)
    constructor_col = next((i for i, h in enumerate(lowered) if "constructor" in h), None)
    chassis_col = next((i for i, h in enumerate(lowered) if "chassis" in h), None)
    driver_col = next((i for i, h in enumerate(lowered) if "driver" in h), None)
    if driver_col is None:
        return []

    records = []
    seen = set()
    for row in matrix[header_rows:]:
        if not row or driver_col >= len(row):
            continue
        drivers = driver_names_from_cell(row[driver_col])
        if not drivers:
            continue
        entrant_cell = row[entrant_col] if entrant_col is not None and entrant_col < len(row) else None
        entrant = clean_wiki_text(text_of(entrant_cell)) if entrant_cell is not None else ""
        constructor = clean_wiki_text(text_of(row[constructor_col])) if constructor_col is not None and constructor_col < len(row) and row[constructor_col] is not None else ""
        chassis = clean_wiki_text(text_of(row[chassis_col])) if chassis_col is not None and chassis_col < len(row) and row[chassis_col] is not None else ""
        entrant_url = link_from_cell(entrant_cell)
        constructor_cell = row[constructor_col] if constructor_col is not None and constructor_col < len(row) else None
        chassis_cell = row[chassis_col] if chassis_col is not None and chassis_col < len(row) else None
        constructor_url = link_from_cell(constructor_cell)
        chassis_url = link_from_cell(chassis_cell)
        for driver in drivers:
            key = (driver_key(driver), norm(entrant), norm(constructor), norm(chassis))
            if key in seen:
                continue
            seen.add(key)
            records.append(
                Entry(
                    driver=driver,
                    driver_key=driver_key(driver),
                    entrant=entrant,
                    entrant_url=entrant_url,
                    constructor=constructor,
                    constructor_url=constructor_url,
                    chassis=chassis,
                    chassis_url=chassis_url,
                )
            )
    return records


def best_entry(row_rec, entries: list[Entry]) -> tuple[Entry | None, float]:
    best = None
    best_score = 0.0
    for entry in entries:
        driver_score = driver_similarity(row_rec["driver"], entry.driver)
        if driver_score < 0.84:
            continue
        team_score = ratio(norm(row_rec["team"]), norm(entry.entrant)) if row_rec["team"] else 0.0
        constructor_score = ratio(norm(row_rec["car"]), norm(entry.constructor))
        chassis_score = ratio(norm(row_rec["type"]), norm(entry.chassis))
        score = driver_score * 0.45 + max(team_score, constructor_score, chassis_score) * 0.55
        if score > best_score:
            best = entry
            best_score = score
    return best, best_score


def best_constructor(row_rec, standings: dict[str, Standing]) -> tuple[Standing | None, float]:
    best = None
    best_score = 0.0
    row_car = norm(row_rec["car"])
    for key, standing in standings.items():
        score = ratio(row_car, key)
        if score > best_score:
            best = standing
            best_score = score
    return best, best_score


ENGINE_TOKENS = {
    "acer",
    "alfa",
    "bmw",
    "cosworth",
    "fondmetal",
    "ford",
    "hart",
    "honda",
    "ilmor",
    "judd",
    "lamborghini",
    "mecachrome",
    "mechachrome",
    "mercedes",
    "mugen",
    "peugeot",
    "playlife",
    "porsche",
    "rbpt",
    "renault",
    "supertec",
    "tag",
    "yamaha",
}


def constructor_root(value: str) -> str:
    tokens = norm(value).split()
    if not tokens:
        return ""
    if len(tokens) >= 2 and tokens[0] == "alfa" and tokens[1] == "romeo":
        return "alfa romeo"
    root = []
    for token in tokens:
        if token in ENGINE_TOKENS and root:
            break
        root.append(token)
    return " ".join(root) if root else " ".join(tokens)


def best_constructor_with_fallback(row_rec, standings: dict[str, Standing], entries: list[Entry]) -> tuple[Standing | None, float]:
    standing, score = best_constructor(row_rec, standings)
    if standing and score >= 0.82:
        return standing, score

    entry, entry_score = best_entry(row_rec, entries)
    if entry and entry_score >= 0.82:
        entry_rec = dict(row_rec)
        entry_rec["car"] = entry.constructor
        standing2, score2 = best_constructor(entry_rec, standings)
        if standing2 and score2 >= 0.82:
            return standing2, score2

    row_root = constructor_root(row_rec["car"])
    if row_root:
        root_matches = [standing for standing in standings.values() if constructor_root(standing.name) == row_root]
        if len(root_matches) == 1:
            return root_matches[0], 0.9
    return standing, score


def set_plain_number_cell(cell, value):
    cell.value = value
    cell._hyperlink = None
    font = copy(cell.font) if cell.font else Font(name="Calibri", size=11)
    font.underline = None
    font.color = "000000"
    cell.font = font


def set_team_cell(cell, value: str, url: str, add_link: bool):
    cell.value = value
    cell._hyperlink = None
    font = copy(cell.font) if cell.font else Font(name="Calibri", size=11)
    if add_link and url:
        cell.hyperlink = url
        font.underline = "single"
        font.color = "0000FF"
    else:
        font.underline = None
        font.color = "000000"
    cell.font = font


def row_record(ws, row: int):
    return {
        "year": int(ws.cell(row, COL_YEAR).value),
        "team": clean(ws.cell(row, COL_TEAM).value),
        "car": clean(ws.cell(row, COL_CAR).value),
        "type": clean(ws.cell(row, COL_TYPE).value),
        "driver": clean(ws.cell(row, COL_DRIVER).value),
        "pc0": pc_is_zero(ws.cell(row, COL_PC).value),
    }


def main():
    wb = load_workbook(INPUT)
    ws = wb[SHEET]

    source_rows = []
    data_by_year = {}
    for year in range(START_YEAR, END_YEAR + 1):
        url, doc = fetch_doc(year)
        drivers = parse_standings_table(doc, "driver")
        constructors = parse_standings_table(doc, "constructor")
        entries = parse_entries_with_links(year, doc)
        data_by_year[year] = {
            "url": url,
            "drivers": drivers,
            "constructors": constructors,
            "entries": entries,
        }
        source_rows.append(
            {
                "year": year,
                "url": url,
                "drivers": len(drivers),
                "constructors": len(constructors),
                "entries": len(entries),
            }
        )
        time.sleep(0.1)

    updated_driver = 0
    updated_constructor = 0
    filled_team = 0
    unmatched_driver = []
    unmatched_constructor = []
    unmatched_team = []

    for row in range(2, ws.max_row + 1):
        year_value = ws.cell(row, COL_YEAR).value
        if not isinstance(year_value, int) or not START_YEAR <= year_value <= END_YEAR:
            continue
        rec = row_record(ws, row)
        year_data = data_by_year[rec["year"]]

        driver_standing, driver_score = best_driver_standing(rec["driver"], year_data["drivers"])
        if driver_standing and driver_score >= 0.84:
            set_plain_number_cell(ws.cell(row, COL_D), driver_standing.pos)
            set_plain_number_cell(ws.cell(row, COL_DP), driver_standing.points)
            updated_driver += 1
        elif clean(ws.cell(row, COL_D).value) not in {"-", ""} or is_url(ws.cell(row, COL_DP).value):
            unmatched_driver.append((row, rec["year"], rec["driver"], rec["car"], rec["type"]))

        constructor_standing, constructor_score = best_constructor_with_fallback(rec, year_data["constructors"], year_data["entries"])
        if constructor_standing and constructor_score >= 0.82:
            set_plain_number_cell(ws.cell(row, COL_T), constructor_standing.pos)
            set_plain_number_cell(ws.cell(row, COL_TP), constructor_standing.points)
            updated_constructor += 1
        elif clean(ws.cell(row, COL_T).value) not in {"-", ""} or is_url(ws.cell(row, COL_TP).value):
            unmatched_constructor.append((row, rec["year"], rec["car"], constructor_score))

        team_cell = ws.cell(row, COL_TEAM)
        if not clean(team_cell.value) or is_url(team_cell.value):
            entry, score = best_entry(rec, year_data["entries"])
            if entry and score >= 0.82:
                set_team_cell(team_cell, entry.entrant, entry.constructor_url or entry.entrant_url or entry.chassis_url, not rec["pc0"])
                filled_team += 1
            else:
                unmatched_team.append((row, rec["year"], rec["driver"], rec["car"], rec["type"], score))

        if clean(ws.cell(row, COL_D).value).upper() == "NC" and clean(ws.cell(row, COL_DP).value) in {"", "-"}:
            set_plain_number_cell(ws.cell(row, COL_DP), 0)
        if clean(ws.cell(row, COL_T).value).upper() == "NC" and clean(ws.cell(row, COL_TP).value) in {"", "-"}:
            set_plain_number_cell(ws.cell(row, COL_TP), 0)
        if rec["pc0"] and clean(ws.cell(row, COL_D).value).upper() == "NC" and clean(ws.cell(row, COL_T).value) == "":
            set_plain_number_cell(ws.cell(row, COL_T), "NC")
            set_plain_number_cell(ws.cell(row, COL_TP), 0)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    with SOURCE_AUDIT.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["year", "url", "drivers", "constructors", "entries"])
        writer.writeheader()
        writer.writerows(source_rows)

    check_wb = load_workbook(OUTPUT)
    check_ws = check_wb[SHEET]
    visible_url_names = []
    pc0_links = 0
    linked_bad_display = []
    for row in range(2, check_ws.max_row + 1):
        year = check_ws.cell(row, COL_YEAR).value
        if not isinstance(year, int) or not START_YEAR <= year <= END_YEAR:
            continue
        pc0 = pc_is_zero(check_ws.cell(row, COL_PC).value)
        for col in [COL_TEAM, COL_DRIVER, COL_D, COL_DP, COL_T, COL_TP]:
            cell = check_ws.cell(row, col)
            if is_url(cell.value):
                visible_url_names.append((row, col, year, cell.value))
            if cell.hyperlink and (cell.value is None or is_url(cell.value)):
                linked_bad_display.append((row, col, year, cell.value, cell.hyperlink.target))
        if pc0:
            for col in range(1, MAX_COL + 1):
                if check_ws.cell(row, col).hyperlink:
                    pc0_links += 1

    print(f"saved_to={OUTPUT}")
    print(f"updated_driver_rows={updated_driver}")
    print(f"updated_constructor_rows={updated_constructor}")
    print(f"filled_blank_team_rows={filled_team}")
    print(f"visible_url_values_in_checked_columns={len(visible_url_names)}")
    print(f"bad_link_display_cells={len(linked_bad_display)}")
    print(f"pc0_links={pc0_links}")
    print(f"unmatched_driver_rows={len(unmatched_driver)}")
    print(f"unmatched_constructor_rows={len(unmatched_constructor)}")
    print(f"unmatched_team_rows={len(unmatched_team)}")
    for label, rows in [
        ("unmatched_driver_sample", unmatched_driver),
        ("unmatched_constructor_sample", unmatched_constructor),
        ("unmatched_team_sample", unmatched_team),
    ]:
        if rows:
            print(label)
            for item in rows[:20]:
                print(item)


if __name__ == "__main__":
    main()
