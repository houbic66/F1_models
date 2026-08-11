from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


INPUT = Path("outputs") / "wiki_audit" / "Diecast 2026 - doplneno z auditu - body 1976-2025.xlsx"


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


wb = load_workbook(INPUT, read_only=True, data_only=False)
print("sheets", wb.sheetnames)
ws = wb["Overview"]
headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
print("max_row", ws.max_row, "max_col", ws.max_column)
print("headers", headers)
for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 8), values_only=True):
    print([clean(v) for v in row[:16]])
