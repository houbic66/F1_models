from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

from catalog_rules import canonical_match_code, infer_manufacturer, is_non_f1_model as rules_is_non_f1_model


INPUT_WORKBOOK = Path("outputs") / "wiki_audit" / "Diecast 2026 - doplneno z auditu - body 1976-2025.xlsx"
RAW_CATALOG = Path("outputs") / "model_catalog" / "sourced_model_catalog_expanded_raw.json"
OUT_DIR = Path("outputs") / "model_catalog"
MATCHED_JSON = OUT_DIR / "all_f1_143_models_matched.json"
MATCHED_CSV = OUT_DIR / "all_f1_143_models_matched.csv"
SUMMARY_JSON = OUT_DIR / "all_f1_143_models_summary.json"


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def deaccent(value: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch))


def norm(value: str) -> str:
    value = deaccent(clean(value)).lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def norm_code(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", deaccent(clean(value)).lower())


def canonical_code(value: str, brand: str = "") -> str:
    return canonical_match_code(value, brand)


def is_non_f1_model(model: dict) -> bool:
    return rules_is_non_f1_model(
        model,
        [
            "constructor_car",
            "chassis_type",
            "driver",
            "race_gp_version",
            "raw_title",
            "notes",
        ],
    )


def driver_tokens(value: str) -> tuple[str, str]:
    value = clean(value)
    if not value:
        return "", ""
    value = value.replace(".", ". ")
    parts = [p for p in re.split(r"\s+", norm(value)) if p and p not in {"jr", "junior"}]
    if not parts:
        return "", ""
    if len(parts[0]) == 1 and len(parts) >= 2:
        return parts[-1], parts[0]
    if len(parts) >= 2:
        # Collection uses "Surname Given", source often "Given Surname".
        return parts[0], parts[1][0]
    return parts[0], ""


def driver_match(a: str, b: str) -> bool:
    if not a or not b:
        return False
    if norm(a) == norm(b):
        return True
    a_last, a_initial = driver_tokens(a)
    b_last, b_initial = driver_tokens(b)
    return bool(a_last and b_last and a_last == b_last and (not a_initial or not b_initial or a_initial == b_initial))


def text_overlap(a: str, b: str) -> float:
    aa = set(norm(a).split())
    bb = set(norm(b).split())
    if not aa or not bb:
        return 0.0
    return len(aa & bb) / max(len(aa), len(bb))


def owned_pc(value: str) -> bool:
    value = clean(value)
    return bool(value) and value not in {"0", "0.0"}


def load_collection() -> list[dict]:
    wb = load_workbook(INPUT_WORKBOOK, read_only=True, data_only=False)
    ws = wb["Overview"]
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item = {
            "row": idx,
            "Year": clean(row[0]),
            "Team": clean(row[1]),
            "Car": clean(row[2]),
            "Type": clean(row[3]),
            "Nr": clean(row[4]),
            "Driver": clean(row[5]),
            "Brand": clean(row[6]),
            "Extra": clean(row[7]),
            "Pc": clean(row[12]),
            "Code": clean(row[13]),
        }
        item["owned"] = owned_pc(item["Pc"])
        rows.append(item)
    return rows


def collection_indexes(rows: list[dict]) -> tuple[dict[str, list[dict]], dict[str, list[dict]]]:
    by_code: dict[str, list[dict]] = defaultdict(list)
    by_year: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        code = canonical_code(row["Code"], row["Brand"])
        if code:
            by_code[code].append(row)
        if row["Year"]:
            by_year[row["Year"]].append(row)
    return by_code, by_year


def find_best_match(model: dict, by_code: dict[str, list[dict]], by_year: dict[str, list[dict]]) -> tuple[str, dict | None, int]:
    code = canonical_code(model.get("model_code", ""), model.get("manufacturer", ""))
    if code and code in by_code:
        candidates = by_code[code]
        candidates = sorted(candidates, key=lambda row: (not row["owned"], row["row"]))
        row = candidates[0]
        return ("Vlastněno - přesný kód" if row["owned"] else "V sešitu, Pc=0 - přesný kód", row, 100)

    year_rows = by_year.get(clean(model.get("year", "")), [])
    best: tuple[int, dict | None] = (0, None)
    for row in year_rows:
        score = 0
        if driver_match(model.get("driver", ""), row["Driver"]):
            score += 45
        if norm(model.get("manufacturer", "")) and norm(model.get("manufacturer", "")) == norm(row["Brand"]):
            score += 20
        car_text = clean(f"{model.get('constructor_car', '')} {model.get('chassis_type', '')}")
        row_text = clean(f"{row['Car']} {row['Type']}")
        overlap = text_overlap(car_text, row_text)
        score += round(overlap * 30)
        if model.get("car_number") and norm(model["car_number"]) in norm(row["Nr"]):
            score += 5
        if score > best[0]:
            best = (score, row)

    score, row = best
    if row and score >= 78:
        return ("Vlastněno - pravděpodobná shoda" if row["owned"] else "V sešitu, Pc=0 - pravděpodobná shoda", row, score)
    if row and score >= 55:
        return ("Možná shoda ve sbírce" if row["owned"] else "Možná shoda s řádkem Pc=0", row, score)
    return "Nenalezeno ve sbírce", None, score


def main() -> None:
    collection = load_collection()
    by_code, by_year = collection_indexes(collection)
    raw_models = json.loads(RAW_CATALOG.read_text(encoding="utf-8"))
    out_rows = []
    for model in raw_models:
        if is_non_f1_model(model):
            continue
        if not clean(model.get("manufacturer")):
            model = {**model, "manufacturer": infer_manufacturer(model.get("manufacturer"), model.get("model_code"), model.get("raw_title"))}
        status, match, score = find_best_match(model, by_code, by_year)
        out_rows.append(
            {
                "Year": model.get("year", ""),
                "Constructor/Car": model.get("constructor_car", ""),
                "Chassis/Type": model.get("chassis_type", ""),
                "Driver": model.get("driver", ""),
                "Car number": model.get("car_number", ""),
                "Team/livery": model.get("team_livery", ""),
                "Race/GP/version": model.get("race_gp_version", ""),
                "Manufacturer": model.get("manufacturer", ""),
                "Model code": model.get("model_code", ""),
                "Scale": model.get("scale", "1/43"),
                "Source URL": model.get("source_url", ""),
                "Match status against collection": status,
                "Collection Pc": match["Pc"] if match else "",
                "Collection row": match["row"] if match else "",
                "Collection Code": match["Code"] if match else "",
                "Collection Driver": match["Driver"] if match else "",
                "Collection Car": match["Car"] if match else "",
                "Collection Type": match["Type"] if match else "",
                "Collection Extra": match["Extra"] if match else "",
                "Match score": score,
                "Source name": model.get("source_name", ""),
                "Raw source title": model.get("raw_title", ""),
                "Limited edition": model.get("limited_edition", ""),
                "Price AUD": model.get("price_aud", ""),
                "Notes": model.get("notes", ""),
            }
        )
    out_rows.sort(key=lambda r: (r["Year"] or "9999", r["Constructor/Car"], r["Chassis/Type"], r["Driver"], r["Manufacturer"], r["Model code"]))
    MATCHED_JSON.write_text(json.dumps(out_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    with MATCHED_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(out_rows[0].keys()))
        writer.writeheader()
        writer.writerows(out_rows)

    summary = {
        "total_models": len(out_rows),
        "by_match_status": dict(Counter(row["Match status against collection"] for row in out_rows)),
        "by_manufacturer": dict(Counter(row["Manufacturer"] for row in out_rows)),
        "by_source": dict(Counter(row["Source name"] for row in out_rows)),
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
