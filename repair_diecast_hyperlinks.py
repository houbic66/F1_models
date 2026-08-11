from __future__ import annotations

import re
from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font


ORIGINAL = Path("originalni_soubor") / "Diecast 2026.xlsx"
TARGET = Path("outputs") / "wiki_audit" / "Diecast 2026 - doplneno z auditu.xlsx"
FALLBACK = Path("outputs") / "wiki_audit" / "Diecast 2026 - doplneno z auditu - opraveno.xlsx"
SHEET = "Overview"

MAX_COL = 16
COL_YEAR = 1
COL_TEAM = 2
COL_CAR = 3
COL_TYPE = 4
COL_NR = 5
COL_DRIVER = 6
COL_BRAND = 7
COL_EXTRA = 8
COL_D = 9
COL_PC = 13
COL_CODE = 14


def clean(value) -> str:
    if value is None:
        return ""
    value = str(value).replace("\xa0", " ")
    return re.sub(r"\s+", " ", value).strip()


def norm(value) -> str:
    value = clean(value).casefold()
    value = re.sub(r"https?://\S+", "", value)
    value = re.sub(r"[^a-z0-9]+", "", value)
    return value


def is_url(value) -> bool:
    return isinstance(value, str) and value.lower().startswith(("http://", "https://"))


def driver_key(value) -> str:
    value = clean(value)
    if is_url(value):
        value = value.rsplit("/", 1)[-1].replace("_", " ")
    parts = re.findall(r"[A-Za-zÀ-ž]+", value.casefold())
    suffixes = {"jr", "sr", "ii", "iii"}
    parts = [p for p in parts if p not in suffixes]
    return "".join(parts)


def pc_is_zero(value) -> bool:
    if value is None:
        return False
    text = clean(value)
    return text == "0" or text == "0.0"


def year_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def row_record(ws, row: int) -> dict:
    return {
        "row": row,
        "year": year_int(ws.cell(row, COL_YEAR).value),
        "team": clean(ws.cell(row, COL_TEAM).value),
        "car": clean(ws.cell(row, COL_CAR).value),
        "type": clean(ws.cell(row, COL_TYPE).value),
        "nr": clean(ws.cell(row, COL_NR).value),
        "driver": clean(ws.cell(row, COL_DRIVER).value),
        "brand": clean(ws.cell(row, COL_BRAND).value),
        "extra": clean(ws.cell(row, COL_EXTRA).value),
        "d": clean(ws.cell(row, COL_D).value),
        "pc": clean(ws.cell(row, COL_PC).value),
        "code": clean(ws.cell(row, COL_CODE).value),
    }


def match_score(out_row: dict, orig_row: dict) -> int:
    score = 0
    if driver_key(out_row["driver"]) and driver_key(out_row["driver"]) == driver_key(orig_row["driver"]):
        score += 4
    if norm(out_row["car"]) and norm(out_row["car"]) == norm(orig_row["car"]):
        score += 3
    if norm(out_row["type"]) and norm(out_row["type"]) == norm(orig_row["type"]):
        score += 3
    if out_row["nr"] and out_row["nr"] == orig_row["nr"]:
        score += 2
    if out_row["d"] and out_row["d"] == orig_row["d"]:
        score += 2
    if out_row["pc"] and out_row["pc"] == orig_row["pc"]:
        score += 2
    if out_row["code"] and out_row["code"] == orig_row["code"]:
        score += 4
    if norm(out_row["brand"]) and norm(out_row["brand"]) == norm(orig_row["brand"]):
        score += 1
    if norm(out_row["extra"]) and norm(out_row["extra"]) == norm(orig_row["extra"]):
        score += 1
    if not is_url(out_row["team"]) and norm(out_row["team"]) and norm(out_row["team"]) == norm(orig_row["team"]):
        score += 2
    return score


def clone_cell(src, dst):
    dst.value = src.value
    if src.has_style:
        dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)
    if src.hyperlink:
        link = copy(src.hyperlink)
        link.ref = dst.coordinate
        dst._hyperlink = link
    else:
        dst._hyperlink = None


def plain_no_link(cell):
    cell._hyperlink = None
    font = copy(cell.font) if cell.font else Font(name="Calibri", size=11)
    font.underline = None
    font.color = "000000"
    cell.font = font


def copy_original_row(orig_ws, out_ws, orig_row: int, out_row: int):
    out_ws.row_dimensions[out_row].height = orig_ws.row_dimensions[orig_row].height
    for col in range(1, MAX_COL + 1):
        clone_cell(orig_ws.cell(orig_row, col), out_ws.cell(out_row, col))


def clear_pc0_links(out_ws):
    cleared = 0
    for row in range(2, out_ws.max_row + 1):
        year = year_int(out_ws.cell(row, COL_YEAR).value)
        if year is None or not 1976 <= year <= 2025:
            continue
        if not pc_is_zero(out_ws.cell(row, COL_PC).value):
            continue
        for col in range(1, MAX_COL + 1):
            cell = out_ws.cell(row, col)
            if cell.hyperlink:
                cleared += 1
            plain_no_link(cell)
    return cleared


def count_visible_url_problems(out_ws):
    problems = []
    checked_cols = [1, 2, 3, 4, 5, 6, 7, 8, 13, 14, 15, 16]
    for row in range(2, out_ws.max_row + 1):
        year = year_int(out_ws.cell(row, COL_YEAR).value)
        if year is None or not 1976 <= year <= 2025:
            continue
        if pc_is_zero(out_ws.cell(row, COL_PC).value):
            continue
        for col in checked_cols:
            if is_url(out_ws.cell(row, col).value):
                problems.append((row, col, out_ws.cell(row, col).value))
    return problems


def main():
    orig_wb = load_workbook(ORIGINAL)
    out_wb = load_workbook(TARGET)
    orig_ws = orig_wb[SHEET]
    out_ws = out_wb[SHEET]

    originals_by_year = defaultdict(list)
    for row in range(2, orig_ws.max_row + 1):
        rec = row_record(orig_ws, row)
        if rec["year"] is not None:
            originals_by_year[rec["year"]].append(rec)

    restored = 0
    unmatched = []
    for row in range(2, out_ws.max_row + 1):
        year = year_int(out_ws.cell(row, COL_YEAR).value)
        if year is None or not 1976 <= year <= 2025:
            continue
        if pc_is_zero(out_ws.cell(row, COL_PC).value):
            continue
        out_rec = row_record(out_ws, row)
        candidates = originals_by_year.get(year, [])
        if not candidates:
            continue
        best = max(candidates, key=lambda item: match_score(out_rec, item))
        score = match_score(out_rec, best)
        if score >= 8:
            copy_original_row(orig_ws, out_ws, best["row"], row)
            restored += 1
        elif any(is_url(out_ws.cell(row, col).value) for col in [COL_TEAM, COL_TYPE, COL_DRIVER]):
            unmatched.append((row, year, score, out_rec))

    cleared = clear_pc0_links(out_ws)

    try:
        out_wb.save(TARGET)
        saved_to = TARGET
    except PermissionError:
        out_wb.save(FALLBACK)
        saved_to = FALLBACK

    check_wb = load_workbook(saved_to)
    check_ws = check_wb[SHEET]
    problems = count_visible_url_problems(check_ws)
    pc0_links = 0
    for row in range(2, check_ws.max_row + 1):
        year = year_int(check_ws.cell(row, COL_YEAR).value)
        if year is None or not 1976 <= year <= 2025:
            continue
        if pc_is_zero(check_ws.cell(row, COL_PC).value):
            for col in range(1, MAX_COL + 1):
                if check_ws.cell(row, col).hyperlink:
                    pc0_links += 1

    print(f"saved_to={saved_to}")
    print(f"restored_rows={restored}")
    print(f"pc0_links_cleared={cleared}")
    print(f"remaining_visible_url_problems={len(problems)}")
    print(f"remaining_pc0_links={pc0_links}")
    if unmatched:
        print("unmatched_problem_rows:")
        for item in unmatched[:20]:
            print(item)


if __name__ == "__main__":
    main()
