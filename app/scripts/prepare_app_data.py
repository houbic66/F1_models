from __future__ import annotations

import json
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
APP_DIR = ROOT / "app"
DATA_DIR = APP_DIR / "data"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from catalog_rules import (  # noqa: E402
    canonical_display_code,
    infer_manufacturer,
    is_minichamps_manufacturer as rules_is_minichamps_manufacturer,
    is_non_f1_model as rules_is_non_f1_model,
    is_spark_manufacturer as rules_is_spark_manufacturer,
)

CATALOG_PATH = ROOT / "outputs" / "model_catalog" / "all_f1_143_models_matched.json"
SUMMARY_PATH = ROOT / "outputs" / "model_catalog" / "all_f1_143_models_summary.json"
PILOT_WORKBOOK = ROOT / "input" / "F1_1980_Driver_Standings_Collection_v10.xlsx"
COLLECTION_WORKBOOK = (
    ROOT
    / "outputs"
    / "wiki_audit"
    / "Diecast 2026 - doplneno z auditu - body 1976-2025.xlsx"
)
PHOTO_OVERRIDES_PATH = DATA_DIR / "model_photo_overrides.json"


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip()
    return re.sub(r"\s+", " ", text)


def deaccent(value: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(char)
    )


def slug(value: str) -> str:
    text = deaccent(clean(value)).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "unknown"


def model_id(manufacturer: str, catalog_number: str) -> str:
    return f"{slug(manufacturer)}__{slug(catalog_number)}"


def compact_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", deaccent(clean(value)).lower())


def canonical_manufacturer_name(value: str) -> str:
    text = clean(value)
    compact = compact_key(text)
    aliases = {
        "sparkjespark": "Spark",
        "jeminichamps": "Minichamps",
        "minichampsjeminichamps": "Minichamps",
    }
    if compact in aliases:
        return aliases[compact]
    if re.search(r"\bspark\b", text, flags=re.I):
        return "Spark"
    if re.search(r"\bminichamps\b", text, flags=re.I):
        return "Minichamps"
    return text


def row_manufacturer_name(value: Any, catalog_number: Any = "", raw_title: Any = "") -> str:
    return canonical_manufacturer_name(infer_manufacturer(value, catalog_number, raw_title))


def is_spark_manufacturer(value: str) -> bool:
    return rules_is_spark_manufacturer(value)


def canonical_catalog_number(manufacturer: str, catalog_number: str) -> str:
    return canonical_display_code(manufacturer, catalog_number)


def inferred_catalog_number(row: dict[str, Any], manufacturer: str, catalog_number: str) -> str:
    direct = canonical_display_code(
        manufacturer,
        catalog_number,
        row.get("Collection Code"),
        row.get("Raw source title"),
        row.get("Source URL"),
    )
    if direct:
        return direct
    if not is_spark_manufacturer(manufacturer):
        if rules_is_minichamps_manufacturer(manufacturer):
            return ""
        return clean(catalog_number)
    return ""


def is_non_f1_catalog_row(row: dict[str, Any]) -> bool:
    return rules_is_non_f1_model(
        row,
        [
            "Constructor/Car",
            "Chassis/Type",
            "Driver",
            "Race/GP/version",
            "Raw source title",
            "Notes",
        ],
    )


def photo_record_for(
    photo_overrides: dict[str, dict[str, Any]],
    manufacturer: str,
    catalog_number: str,
    fallback_id: str = "",
) -> dict[str, Any]:
    if fallback_id and fallback_id in photo_overrides:
        return photo_overrides[fallback_id]

    manufacturer_slug = slug(manufacturer)
    code_slug = slug(catalog_number)
    keys = [f"{manufacturer_slug}__{code_slug}"]

    code_compact = compact_key(catalog_number)
    if manufacturer_slug == "spark" and re.fullmatch(r"s\d+", code_compact):
        keys.append(f"spark__spk{code_compact[1:]}")
    if manufacturer_slug == "spark" and re.fullmatch(r"spk\d+", code_compact):
        keys.append(f"spark__s{code_compact[3:]}")
    raceland = re.search(r"rs(\d{4})", code_compact)
    if manufacturer_slug == "spark" and raceland:
        keys.append(f"spark__rs{raceland.group(1)}")

    for key in keys:
        if key in photo_overrides:
            return photo_overrides[key]

    wanted = f"{manufacturer_slug}__{code_compact}"
    for key, record in photo_overrides.items():
        if compact_key(key) == compact_key(wanted):
            return record

    return normalize_photo_record(None)


def number(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def source_urls(value: str) -> list[str]:
    return [part.strip() for part in clean(value).split(" | ") if part.strip()]


def normalize_photo_record(value: Any) -> dict[str, Any]:
    if isinstance(value, list):
        urls = [clean(url) for url in value if clean(url)]
        return {
            "mainPhoto": urls[0] if urls else "",
            "thumbnails": urls[1:],
            "originalPhotoUrl": urls[0] if urls else "",
            "sourcePageUrl": "",
            "photoStatus": "verified" if urls else "missing",
            "photoCheckedAt": "",
            "photoHttpStatus": "",
            "photoContentType": "",
        }
    if isinstance(value, dict):
        thumbnails = value.get("thumbnails", [])
        if isinstance(thumbnails, str):
            thumbnails = [thumbnails]
        thumbnails = [clean(url) for url in thumbnails if clean(url)]
        main = clean(value.get("mainPhoto") or value.get("main") or (thumbnails[0] if thumbnails else ""))
        original = clean(value.get("originalPhotoUrl") or value.get("original") or main)
        return {
            "mainPhoto": main,
            "thumbnails": thumbnails,
            "originalPhotoUrl": original,
            "sourcePageUrl": clean(value.get("sourcePageUrl") or value.get("source") or ""),
            "photoStatus": clean(value.get("photoStatus")) or ("verified" if main else "missing"),
            "photoCheckedAt": clean(value.get("photoCheckedAt")),
            "photoHttpStatus": clean(value.get("photoHttpStatus")),
            "photoContentType": clean(value.get("photoContentType")),
        }
    return {
        "mainPhoto": "",
        "thumbnails": [],
        "originalPhotoUrl": "",
        "sourcePageUrl": "",
        "photoStatus": "missing",
        "photoCheckedAt": "",
        "photoHttpStatus": "",
        "photoContentType": "",
    }


def load_photo_overrides() -> dict[str, dict[str, Any]]:
    if not PHOTO_OVERRIDES_PATH.exists():
        return {}
    raw = json.loads(PHOTO_OVERRIDES_PATH.read_text(encoding="utf-8"))
    return {clean(key): normalize_photo_record(value) for key, value in raw.items()}


def photo_urls(photo_record: dict[str, Any]) -> list[str]:
    return [url for url in [photo_record.get("mainPhoto", ""), *photo_record.get("thumbnails", [])] if clean(url)]


def collection_display_status(quantity: int, nv: int, v: int, code: str) -> str:
    if clean(code).upper() == "NO MODEL":
        return "red"
    if v > 0:
        return "green"
    if quantity > 0 or nv > 0:
        return "white"
    return "yellow"


def collection_display_label(status: str) -> str:
    return {
        "green": "Ve vitríně",
        "white": "Mimo vitrínu",
        "yellow": "Chybí",
        "red": "NO MODEL",
    }.get(status, "Chybí")


def row_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}


def load_collection(photo_overrides: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    wb = load_workbook(COLLECTION_WORKBOOK, read_only=True, data_only=True)
    ws = wb["Overview"]
    headers = [clean(cell.value) for cell in ws[1]]
    items: list[dict[str, Any]] = []
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = row_dict(headers, row)
        raw_code = clean(values.get("Code"))
        manufacturer = row_manufacturer_name(values.get("Brand"), raw_code, values.get("Extra"))
        code = canonical_catalog_number(manufacturer, raw_code) or raw_code
        quantity = number(values.get("Pc"))
        nv = number(values.get("NV"))
        v = number(values.get("V"))
        display_status = collection_display_status(quantity, nv, v, code)
        item_id = model_id(manufacturer, code) if code and manufacturer else f"collection-row-{row_index}"
        photo_record = photo_record_for(photo_overrides, manufacturer, code, item_id)
        items.append(
            {
                "id": item_id,
                "sourceRow": row_index,
                "season": clean(values.get("Year")),
                "team": clean(values.get("Team")),
                "car": clean(values.get("Car")),
                "chassis": clean(values.get("Type")),
                "carNumber": clean(values.get("Nr")).replace("Nr ", ""),
                "driver": clean(values.get("Driver")),
                "manufacturer": manufacturer,
                "catalogNumber": code,
                "extra": clean(values.get("Extra")),
                "quantity": quantity,
                "nv": nv,
                "v": v,
                "driverStanding": clean(values.get("D")),
                "driverPoints": clean(values.get("DP")),
                "teamStanding": clean(values.get("T")),
                "teamPoints": clean(values.get("TP")),
                "owned": quantity > 0,
                "displayStatus": display_status,
                "displayLabel": collection_display_label(display_status),
                "mainPhoto": photo_record["mainPhoto"],
                "thumbnails": photo_record["thumbnails"],
                "originalPhotoUrl": photo_record["originalPhotoUrl"],
                "photoSourcePageUrl": photo_record["sourcePageUrl"],
                "photoStatus": photo_record["photoStatus"],
                "photoCheckedAt": photo_record["photoCheckedAt"],
                "photoHttpStatus": photo_record["photoHttpStatus"],
                "photoContentType": photo_record["photoContentType"],
                "photoUrls": photo_urls(photo_record),
            }
        )
    return items


def load_pilot_workbook() -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(PILOT_WORKBOOK, read_only=True, data_only=True)
    output: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in ["1980 Driver Order", "Master Index", "Source Registry", "Coverage & Gaps"]:
        ws = wb[sheet_name]
        headers = [clean(cell.value) for cell in ws[1]]
        rows: list[dict[str, Any]] = []
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            values = row_dict(headers, row)
            values["_row"] = row_index
            rows.append({key: clean(value) if not isinstance(value, (int, float)) else value for key, value in values.items()})
        output[sheet_name] = rows
    return output


def derive_color_status(match_status: str, collection_status: str = "", model_status: str = "") -> str:
    text = " ".join([match_status, collection_status, model_status]).lower()
    if "no model" in text:
        return "red"
    if "owned" in text or "vlastněno" in text:
        return "green"
    if "other model owned" in text or "možná" in text:
        return "white"
    if "catalog_only" in text or "model found" in text or "nenalezeno" in text:
        return "yellow"
    return "yellow"


def merge_unique(values: list[str], additions: list[str]) -> list[str]:
    output = [value for value in values if clean(value)]
    for value in additions:
        value = clean(value)
        if value and value not in output:
            output.append(value)
    return output


def token_overlap(a: str, b: str) -> float:
    aa = set(re.findall(r"[a-z0-9]+", deaccent(clean(a)).lower()))
    bb = set(re.findall(r"[a-z0-9]+", deaccent(clean(b)).lower()))
    if not aa or not bb:
        return 0.0
    return len(aa & bb) / len(aa | bb)


def model_match_text(value: dict[str, Any]) -> str:
    return " ".join(
        [
            value.get("constructor", ""),
            value.get("chassis", ""),
            value.get("driver", ""),
            value.get("event", ""),
            value.get("rawTitle", ""),
        ]
    )


def match_car_number(value: dict[str, Any]) -> str:
    direct = clean(value.get("carNumber"))
    if direct:
        return direct
    text = clean(value.get("rawTitle"))
    match = re.search(r"#\s*([0-9]{1,3}[A-Z]?)\b", text, flags=re.I)
    if match:
        return match.group(1)
    return ""


def chassis_match_key(value: dict[str, Any]) -> str:
    chassis = clean(value.get("chassis"))
    match = re.search(r"[A-Z]*\d+[A-Z]*", chassis, flags=re.I)
    return deaccent(match.group(0)).lower() if match else deaccent(chassis.split(" ", 1)[0]).lower()


def duplicates_existing_model(base: dict[str, Any], model_texts: dict[tuple[str, str], list[str]]) -> bool:
    candidate_text = " ".join(
        [
            base.get("constructor", ""),
            base.get("chassis", ""),
            base.get("driver", ""),
            base.get("event", ""),
            base.get("rawTitle", ""),
        ]
    )
    for model_text in model_texts.get((base.get("season", ""), base.get("manufacturer", "")), []):
        if token_overlap(candidate_text, model_text) >= 0.42:
            return True
    return False


def duplicates_coded_spark_model(base: dict[str, Any], models: list[dict[str, Any]]) -> bool:
    if slug(base.get("manufacturer")) != "spark" or clean(base.get("catalogNumber")):
        return False
    base_number = match_car_number(base)
    base_chassis = chassis_match_key(base)
    base_constructor = deaccent(clean(base.get("constructor"))).lower()
    if not base_number or not base_chassis or not base_constructor:
        return False
    for model in models:
        if model.get("season") != base.get("season") or slug(model.get("manufacturer")) != "spark":
            continue
        if not clean(model.get("catalogNumber")):
            continue
        if match_car_number(model) != base_number:
            continue
        if chassis_match_key(model) != base_chassis:
            continue
        if deaccent(clean(model.get("constructor"))).lower() == base_constructor:
            return True
    return False


def driver_last_name(value: str) -> str:
    parts = re.findall(r"[a-z]+", deaccent(clean(value)).lower())
    if not parts:
        return ""
    return parts[0] if len(parts) > 1 and len(parts[0]) != 1 else parts[-1]


def model_has_driver(model: dict[str, Any], driver: str) -> bool:
    last = driver_last_name(driver)
    if not last:
        return False
    text = deaccent(
        " ".join(
            [
                model.get("driver", ""),
                model.get("collectionDriver", ""),
                model.get("event", ""),
                model.get("rawTitle", ""),
                model.get("title", ""),
            ]
        )
    ).lower()
    return last in text


def copy_model_photo_to_collection(item: dict[str, Any], model: dict[str, Any]) -> None:
    item["mainPhoto"] = model.get("mainPhoto", "")
    item["thumbnails"] = model.get("thumbnails", [])
    item["originalPhotoUrl"] = model.get("originalPhotoUrl") or model.get("mainPhoto", "")
    item["photoSourcePageUrl"] = model.get("photoSourcePageUrl", "")
    item["photoStatus"] = model.get("photoStatus", "missing")
    item["photoCheckedAt"] = model.get("photoCheckedAt", "")
    item["photoHttpStatus"] = model.get("photoHttpStatus", "")
    item["photoContentType"] = model.get("photoContentType", "")
    item["photoUrls"] = [url for url in [item["mainPhoto"], *item["thumbnails"]] if clean(url)]


def enrich_collection_photos(collection: list[dict[str, Any]], models: list[dict[str, Any]]) -> None:
    by_id = {model.get("id"): model for model in models}
    for item in collection:
        if item.get("mainPhoto"):
            continue
        exact = by_id.get(item.get("id"))
        if exact and exact.get("mainPhoto"):
            copy_model_photo_to_collection(item, exact)
            continue

        item_maker = clean(item.get("manufacturer"))
        item_chassis = chassis_match_key({"chassis": item.get("chassis", "")})
        item_car_text = clean(f"{item.get('car', '')} {item.get('chassis', '')}")
        best: tuple[int, dict[str, Any] | None] = (0, None)
        for model in models:
            if model.get("season") != item.get("season") or not model.get("mainPhoto"):
                continue
            if item_maker and item_maker != model.get("manufacturer"):
                continue
            if item_chassis and item_chassis != chassis_match_key(model):
                continue
            if not model_has_driver(model, item.get("driver", "")):
                continue
            score = 50 + round(40 * token_overlap(item_car_text, model_match_text(model)))
            if item.get("extra") and token_overlap(item.get("extra", ""), model.get("event", "")) > 0:
                score += 10
            if score > best[0]:
                best = (score, model)
        score, model = best
        if model and score >= 55:
            copy_model_photo_to_collection(item, model)


def match_status_rank(value: str) -> int:
    text = clean(value).lower()
    if "přesn" in text or "presn" in deaccent(text):
        return 4
    if "pravd" in text or "owned" in text:
        return 3
    if "možn" in text or "mozn" in deaccent(text):
        return 2
    if "nenalezeno" in text:
        return 1
    return 0


def merge_model(existing: dict[str, Any], incoming: dict[str, Any]) -> None:
    existing["sourceUrls"] = merge_unique(existing.get("sourceUrls", []), incoming.get("sourceUrls", []))
    existing["photoUrls"] = merge_unique(existing.get("photoUrls", []), incoming.get("photoUrls", []))
    existing["thumbnails"] = merge_unique(existing.get("thumbnails", []), incoming.get("thumbnails", []))
    if incoming.get("photoSourcePageUrl") and incoming["photoSourcePageUrl"] not in existing["sourceUrls"]:
        existing["sourceUrls"].append(incoming["photoSourcePageUrl"])
    if not existing.get("mainPhoto") and incoming.get("mainPhoto"):
        existing["mainPhoto"] = incoming["mainPhoto"]
        existing["originalPhotoUrl"] = incoming.get("originalPhotoUrl", incoming["mainPhoto"])
        existing["photoSourcePageUrl"] = incoming.get("photoSourcePageUrl", "")
    if incoming.get("sourceName") and incoming["sourceName"] not in existing.get("sourceName", ""):
        existing["sourceName"] = " | ".join([part for part in [existing.get("sourceName", ""), incoming["sourceName"]] if part])
    if incoming.get("rawTitle") and incoming["rawTitle"] not in existing.get("rawTitle", ""):
        existing["rawTitle"] = " | ".join([part for part in [existing.get("rawTitle", ""), incoming["rawTitle"]] if part])
    if match_status_rank(incoming.get("matchStatus", "")) > match_status_rank(existing.get("matchStatus", "")):
        for key in [
            "owned",
            "colorStatus",
            "matchStatus",
            "collectionQuantity",
            "collectionRow",
            "collectionCode",
            "collectionDriver",
            "collectionCar",
            "collectionChassis",
            "collectionExtra",
            "matchScore",
        ]:
            existing[key] = incoming[key]


def build_catalog(photo_overrides: dict[str, dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    raw_rows = json.loads(CATALOG_PATH.read_text(encoding="utf-8"))

    def row_priority(row: dict[str, Any]) -> tuple[int, int, int]:
        raw_code = clean(row.get("Model code"))
        manufacturer = row_manufacturer_name(row.get("Manufacturer"), raw_code, row.get("Raw source title"))
        canonical_code = canonical_catalog_number(manufacturer, raw_code)
        has_collection = 0 if clean(row.get("Collection row")) or number(row.get("Collection Pc")) > 0 else 1
        raw_compact = compact_key(raw_code)
        canonical_source = 0 if raw_compact.startswith(("s", "spk")) else 1
        valid_code = 0 if canonical_code else 1
        return (valid_code, has_collection, canonical_source)

    raw_rows = sorted(raw_rows, key=row_priority)
    models: list[dict[str, Any]] = []
    candidates: list[dict[str, Any]] = []
    seen: set[str] = set()
    model_texts: dict[tuple[str, str], list[str]] = {}
    for index, row in enumerate(raw_rows, start=1):
        if is_non_f1_catalog_row(row):
            continue
        raw_catalog_number = clean(row.get("Model code"))
        manufacturer = row_manufacturer_name(row.get("Manufacturer"), raw_catalog_number, row.get("Raw source title"))
        catalog_number = inferred_catalog_number(row, manufacturer, raw_catalog_number)
        base = {
            "season": clean(row.get("Year")),
            "constructor": clean(row.get("Constructor/Car")),
            "chassis": clean(row.get("Chassis/Type")),
            "driver": clean(row.get("Driver")),
            "carNumber": clean(row.get("Car number")),
            "teamLivery": clean(row.get("Team/livery")),
            "event": clean(row.get("Race/GP/version")),
            "manufacturer": manufacturer,
            "catalogNumber": catalog_number,
            "sourceUrls": source_urls(row.get("Source URL", "")),
            "sourceName": clean(row.get("Source name")),
            "rawTitle": clean(row.get("Raw source title")),
            "matchStatus": clean(row.get("Match status against collection")),
            "collectionQuantity": number(row.get("Collection Pc")),
            "collectionRow": clean(row.get("Collection row")),
            "collectionCode": clean(row.get("Collection Code")),
            "collectionDriver": clean(row.get("Collection Driver")),
            "collectionCar": clean(row.get("Collection Car")),
            "collectionChassis": clean(row.get("Collection Type")),
            "collectionExtra": clean(row.get("Collection Extra")),
            "matchScore": number(row.get("Match score")),
            "limitedEdition": clean(row.get("Limited edition")),
            "priceAud": clean(row.get("Price AUD")),
            "notes": clean(row.get("Notes")),
        }
        if manufacturer and catalog_number:
            mid = model_id(manufacturer, catalog_number)
            if mid in seen:
                existing = next(model for model in models if model["id"] == mid)
                duplicate_photo_record = photo_record_for(photo_overrides, manufacturer, catalog_number, mid)
                existing["sourceUrls"] = merge_unique(existing.get("sourceUrls", []), base["sourceUrls"])
                if duplicate_photo_record["sourcePageUrl"]:
                    existing["sourceUrls"] = merge_unique(existing["sourceUrls"], [duplicate_photo_record["sourcePageUrl"]])
                existing["photoUrls"] = merge_unique(existing.get("photoUrls", []), photo_urls(duplicate_photo_record))
                existing["thumbnails"] = merge_unique(existing.get("thumbnails", []), duplicate_photo_record["thumbnails"])
                if not existing.get("mainPhoto") and duplicate_photo_record["mainPhoto"]:
                    existing["mainPhoto"] = duplicate_photo_record["mainPhoto"]
                    existing["originalPhotoUrl"] = duplicate_photo_record["originalPhotoUrl"]
                    existing["photoSourcePageUrl"] = duplicate_photo_record["sourcePageUrl"]
                    existing["photoStatus"] = duplicate_photo_record["photoStatus"]
                    existing["photoCheckedAt"] = duplicate_photo_record["photoCheckedAt"]
                    existing["photoHttpStatus"] = duplicate_photo_record["photoHttpStatus"]
                    existing["photoContentType"] = duplicate_photo_record["photoContentType"]
                if base["sourceName"] and base["sourceName"] not in existing.get("sourceName", ""):
                    existing["sourceName"] = " | ".join([part for part in [existing.get("sourceName", ""), base["sourceName"]] if part])
                continue
            seen.add(mid)
            photo_record = photo_record_for(photo_overrides, manufacturer, catalog_number, mid)
            model_source_urls = merge_unique(base["sourceUrls"], [photo_record["sourcePageUrl"]])
            models.append(
                {
                    "id": mid,
                    "title": " ".join(
                        part
                        for part in [
                            base["season"],
                            base["manufacturer"],
                            base["constructor"],
                            base["chassis"],
                            base["driver"],
                            f"#{base['carNumber']}" if base["carNumber"] else "",
                        ]
                        if part
                    ),
                    "owned": base["collectionQuantity"] > 0 or "Vlastněno" in base["matchStatus"],
                    "colorStatus": derive_color_status(base["matchStatus"]),
                    "mainPhoto": photo_record["mainPhoto"],
                    "thumbnails": photo_record["thumbnails"],
                    "originalPhotoUrl": photo_record["originalPhotoUrl"],
                    "photoSourcePageUrl": photo_record["sourcePageUrl"],
                    "photoStatus": photo_record["photoStatus"],
                    "photoCheckedAt": photo_record["photoCheckedAt"],
                    "photoHttpStatus": photo_record["photoHttpStatus"],
                    "photoContentType": photo_record["photoContentType"],
                    "photoUrls": photo_urls(photo_record),
                    **base,
                    "sourceUrls": model_source_urls,
                }
            )
            model_texts.setdefault((base["season"], base["manufacturer"]), []).append(model_match_text(base))
        else:
            reason = "Spark bez platného SXXXX kódu" if slug(manufacturer) == "spark" else "Chybí výrobce nebo katalogové číslo"
            if duplicates_existing_model(base, model_texts) or duplicates_coded_spark_model(base, models):
                continue
            candidates.append(
                {
                    "id": f"candidate-{index}",
                    "reason": "Chybí výrobce nebo katalogové číslo",
                    "verificationStatus": "NEEDS_REVIEW",
                    "rawCatalogNumber": raw_catalog_number,
                    "reason": reason,
                    **base,
                }
            )
    return models, candidates


def build_app_data() -> dict[str, Any]:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    photos = load_photo_overrides()
    catalog_models, candidates = build_catalog(photos)
    collection = load_collection(photos)
    enrich_collection_photos(collection, catalog_models)
    pilot = load_pilot_workbook()
    summary = json.loads(SUMMARY_PATH.read_text(encoding="utf-8"))
    owned_collection = [item for item in collection if item["owned"]]
    seasons = sorted(
        {
            item["season"]
            for item in catalog_models
            if re.fullmatch(r"(?:19|20)\d{2}", item["season"])
        },
        key=int,
    )
    manufacturers: dict[str, int] = {}
    for model in catalog_models:
        manufacturers[model["manufacturer"]] = manufacturers.get(model["manufacturer"], 0) + 1
    return {
        "generatedAt": date.today().isoformat(),
        "rules": {
            "scale": "1:43",
            "masterKey": "manufacturer + catalogNumber",
            "noCatalogNumberPolicy": "candidate_queue",
            "sourceOfTruth": "Master Catalog describes what exists; Collection describes what is owned.",
            "codeRules": {
                "Spark": "only S + 4 digits is accepted as the canonical model code; seller codes are used only to infer the S code",
                "Minichamps": "only a 9 digit code is accepted as the canonical model code",
                "Other manufacturers": "their own catalog code is preserved until a maker-specific rule is added",
            },
            "excludedSeries": [
                "Formula 2",
                "Formula 3",
                "Formula Ford",
                "IndyCar/CART/Indy 500",
                "Le Mans/sportscar/GT/DTM/touring cars",
            ],
        },
        "summary": {
            "masterModels": len(catalog_models),
            "candidates": len(candidates),
            "collectionRows": len(collection),
            "ownedCollectionRows": len(owned_collection),
            "ownedUniqueCatalogNumbers": len({item["id"] for item in owned_collection if item["catalogNumber"]}),
            "sourceRowsBeforeStrictCatalogFilter": summary.get("total_models"),
            "matchedExactOwned": summary.get("by_match_status", {}).get("Vlastněno - přesný kód", 0),
            "matchedLikelyOwned": summary.get("by_match_status", {}).get("Vlastněno - pravděpodobná shoda", 0),
            "possibleMatches": summary.get("by_match_status", {}).get("Možná shoda ve sbírce", 0),
        },
        "seasons": seasons,
        "manufacturers": sorted(
            [{"name": key, "count": value} for key, value in manufacturers.items()],
            key=lambda item: (-item["count"], item["name"]),
        ),
        "models": catalog_models,
        "candidates": candidates,
        "collectionItems": collection,
        "pilot1980": {
            "seasonRows": pilot["1980 Driver Order"],
            "masterIndex": pilot["Master Index"],
            "sourceRegistry": pilot["Source Registry"],
            "coverageGaps": pilot["Coverage & Gaps"],
        },
    }


def main() -> None:
    app_data = build_app_data()
    (DATA_DIR / "app-data.json").write_text(
        json.dumps(app_data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    if not PHOTO_OVERRIDES_PATH.exists():
        PHOTO_OVERRIDES_PATH.write_text("{}\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "models": app_data["summary"]["masterModels"],
                "candidates": app_data["summary"]["candidates"],
                "collectionRows": app_data["summary"]["collectionRows"],
                "ownedRows": app_data["summary"]["ownedCollectionRows"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
