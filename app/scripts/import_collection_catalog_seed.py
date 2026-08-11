from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
COLLECTION_WORKBOOK = (
    ROOT
    / "outputs"
    / "wiki_audit"
    / "Diecast 2026 - doplneno z auditu - body 1976-2025.xlsx"
)
RAW_CATALOG = ROOT / "outputs" / "model_catalog" / "sourced_model_catalog_expanded_raw.json"


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def number(value: Any) -> int:
    text = clean(value)
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def canonical_manufacturer(value: str) -> str:
    text = clean(value)
    if re.search(r"\bspark\b", text, flags=re.I):
        return "Spark"
    if re.search(r"\bminichamps\b", text, flags=re.I):
        return "Minichamps"
    return text


def load_rows(season: str) -> list[dict[str, str]]:
    wb = load_workbook(COLLECTION_WORKBOOK, read_only=True, data_only=True)
    ws = wb["Overview"]
    headers = [clean(cell.value) for cell in ws[1]]
    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = dict(zip(headers, row))
        year = clean(values.get("Year"))
        if year != season:
            continue
        manufacturer = canonical_manufacturer(values.get("Brand"))
        code = clean(values.get("Code"))
        if not manufacturer or not code or code.lower() == "není":
            continue
        if number(values.get("Pc")) <= 0:
            continue
        rows.append(
            {
                "year": year,
                "constructor_car": clean(values.get("Car")),
                "chassis_type": clean(values.get("Type")),
                "driver": clean(values.get("Driver")),
                "car_number": clean(values.get("Nr")),
                "team_livery": clean(values.get("Team")),
                "race_gp_version": clean(values.get("Extra")),
                "manufacturer": manufacturer,
                "model_code": code,
                "scale": "1/43",
                "source_url": "",
                "source_name": "User collection workbook seed",
                "raw_title": clean(
                    " ".join(
                        [
                            year,
                            manufacturer,
                            clean(values.get("Car")),
                            clean(values.get("Type")),
                            clean(values.get("Driver")),
                            clean(values.get("Nr")),
                            clean(values.get("Extra")),
                            code,
                        ]
                    )
                ),
                "limited_edition": "",
                "price_aud": "",
                "notes": "Seeded from owned collection row with model code",
            }
        )
    return rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--season", default="1981")
    args = parser.parse_args()
    raw = json.loads(RAW_CATALOG.read_text(encoding="utf-8"))
    existing = {
        (
            clean(row.get("year")),
            clean(row.get("manufacturer")).lower(),
            clean(row.get("model_code")).lower(),
        )
        for row in raw
    }
    added = 0
    for row in load_rows(args.season):
        key = (row["year"], row["manufacturer"].lower(), row["model_code"].lower())
        if key in existing:
            continue
        raw.append(row)
        existing.add(key)
        added += 1
    RAW_CATALOG.write_text(json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"season": args.season, "collectionSeedRowsAdded": added}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
