from __future__ import annotations

import csv
import re
import time
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

from audit_f1_wiki import (
    START_YEAR,
    END_YEAR,
    driver_key,
    driver_similarity,
    fetch_doc,
    norm,
    parse_driver_order,
    parse_entries,
    ratio,
)


INPUT = Path("originalni_soubor") / "Diecast 2026.xlsx"
AUDIT_DIR = Path("outputs") / "wiki_audit"
MISSING_CSV = AUDIT_DIR / "missing_in_sheet.csv"
OUTPUT = AUDIT_DIR / "Diecast 2026 - doplneno z auditu.xlsx"

COL_YEAR = 1
COL_TEAM = 2
COL_CAR = 3
COL_TYPE = 4
COL_NR = 5
COL_DRIVER = 6
COL_D = 9
COL_PC = 13
COL_CODE = 14
COL_NV = 15
COL_V = 16
MAX_COL = 16


def clean(value):
    if value is None:
        return ""
    value = str(value).replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value).strip()
    return value


def excel_constructor(value: str) -> str:
    value = clean(value)
    value = value.replace(" - ", "-").replace(" – ", "-")
    return value


def excel_chassis(value: str) -> str:
    value = clean(value)
    return value.split()[0] if " " in value else value


def wiki_to_excel_driver(name: str, preferred: dict[str, str]) -> str:
    key = driver_key(name)
    if key in preferred:
        return preferred[key]
    parts = clean(name).split()
    if not parts:
        return clean(name)
    suffixes = {"Jr.", "Jr", "Sr.", "Sr", "II", "III"}
    parts = [p for p in parts if p not in suffixes]
    particles = {"de", "de la", "van", "von", "di", "da", "del", "della"}
    if len(parts) >= 3 and " ".join(p.lower() for p in parts[-2:]) in particles:
        surname = " ".join(parts[-2:])
        given = " ".join(parts[:-2])
    elif len(parts) >= 2 and parts[-2].lower() in particles:
        surname = " ".join(parts[-2:])
        given = " ".join(parts[:-2])
    else:
        surname = parts[-1]
        given = " ".join(parts[:-1])
    return f"{surname} {given}".strip()


def no_value(value: str) -> str | None:
    value = clean(value)
    if not value:
        return None
    numbers = re.findall(r"\d+", value)
    if len(numbers) == 1:
        return f"Nr {numbers[0]}"
    return f"Nr {value}" if value else None


def d_sort(value):
    if value is None or value == "":
        return (98, 0)
    text = str(value).strip()
    if text.isdigit():
        return (0, int(text))
    if text.upper() == "NC":
        return (90, 0)
    return (95, 0)


def row_match_score(row, ref):
    team_score = ratio(norm(ref["wiki_entrant"]), norm(row["team"]))
    car_score = ratio(norm(ref["wiki_constructor"]), norm(row["car"]))
    type_score = ratio(norm(ref["wiki_chassis"]), norm(row["type"]))
    return max(team_score, car_score, type_score)


def sheet_row_record(ws, row):
    return {
        "row": row,
        "year": ws.cell(row, COL_YEAR).value,
        "team": clean(ws.cell(row, COL_TEAM).value),
        "car": clean(ws.cell(row, COL_CAR).value),
        "type": clean(ws.cell(row, COL_TYPE).value),
        "driver": clean(ws.cell(row, COL_DRIVER).value),
        "pc": ws.cell(row, COL_PC).value,
        "d": ws.cell(row, COL_D).value,
    }


def copy_row_format(ws, src_row, dst_row):
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, MAX_COL + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)


def plain_no_link_style(cell):
    cell.hyperlink = None
    cell._hyperlink = None
    font = copy(cell.font) if cell.font else Font(name="Calibri", size=11)
    font.underline = None
    font.color = "000000"
    cell.font = font


def clear_row_links(ws, row):
    for col in range(1, MAX_COL + 1):
        plain_no_link_style(ws.cell(row, col))


def write_missing_row(ws, row, ref, preferred_drivers, d_value):
    values = {
        COL_YEAR: int(ref["year"]),
        COL_TEAM: clean(ref["wiki_entrant"]),
        COL_CAR: excel_constructor(ref["wiki_constructor"]),
        COL_TYPE: excel_chassis(ref["wiki_chassis"]),
        COL_NR: no_value(ref.get("wiki_no", "")),
        COL_DRIVER: wiki_to_excel_driver(ref["wiki_driver"], preferred_drivers),
        COL_D: d_value,
        COL_PC: 0,
        COL_CODE: None,
        COL_NV: None,
        COL_V: None,
    }
    for col in range(1, MAX_COL + 1):
        cell = ws.cell(row, col)
        cell.value = values.get(col, None)
        cell.hyperlink = None
    clear_row_links(ws, row)


def year_rows(ws, year):
    return [r for r in range(2, ws.max_row + 1) if ws.cell(r, COL_YEAR).value == year]


def find_insert_row(ws, year, driver_key_value, d_value):
    rows = year_rows(ws, year)
    if not rows:
        return ws.max_row + 1
    same_driver = [
        r
        for r in rows
        if driver_key(clean(ws.cell(r, COL_DRIVER).value)) == driver_key_value
    ]
    if same_driver:
        return max(same_driver) + 1
    target_sort = d_sort(d_value)
    for r in rows:
        row_sort = d_sort(ws.cell(r, COL_D).value)
        if row_sort > target_sort:
            return r
    return max(rows) + 1


def build_reference():
    refs = {}
    order = {}
    for year in range(START_YEAR, END_YEAR + 1):
        _, doc = fetch_doc(year)
        year_refs = parse_entries(year, doc)
        for ref in year_refs:
            key = (
                year,
                ref["driver_key"],
                norm(ref["wiki_entrant"]),
                norm(ref["wiki_constructor"]),
                norm(ref["wiki_chassis"]),
            )
            refs[key] = ref
        order[year] = parse_driver_order(year, doc)
        time.sleep(0.1)
    return refs, order


def main():
    wb = load_workbook(INPUT)
    ws = wb["Overview"]

    preferred_counter = defaultdict(Counter)
    for r in range(2, ws.max_row + 1):
        driver = clean(ws.cell(r, COL_DRIVER).value)
        if driver:
            preferred_counter[driver_key(driver)][driver] += 1
    preferred_drivers = {k: counts.most_common(1)[0][0] for k, counts in preferred_counter.items()}

    refs_by_key, order_by_year = build_reference()
    with MISSING_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        missing_rows = list(csv.DictReader(f))

    updated_rows = []
    inserted_rows = []
    skipped_rows = []
    used_update_targets = set()
    resolved_missing = []

    for item in missing_rows:
        year = int(item["year"])
        key = (
            year,
            item["driver_key"],
            norm(item["wiki_entrant"]),
            norm(item["wiki_constructor"]),
            norm(item["wiki_chassis"]),
        )
        ref = refs_by_key.get(key) or {
            "year": year,
            "driver_key": item["driver_key"],
            "wiki_driver": item["wiki_driver"],
            "wiki_entrant": item["wiki_entrant"],
            "wiki_constructor": item["wiki_constructor"],
            "wiki_chassis": item["wiki_chassis"],
            "wiki_no": "",
        }
        order = order_by_year.get(year, {})
        d_value = order.get(ref["driver_key"], {}).get("wiki_pos") or "NC"
        resolved_missing.append((ref, d_value))

        candidates = []
        for r in year_rows(ws, year):
            if r in used_update_targets:
                continue
            row = sheet_row_record(ws, r)
            if not row["driver"] or driver_similarity(row["driver"], ref["wiki_driver"]) < 0.84:
                continue
            pc = row["pc"]
            is_no_model_row = pc in (0, "0", None, "")
            if not is_no_model_row:
                continue
            score = row_match_score(row, ref)
            blanks = sum(1 for field in ("team", "car", "type") if not row[field])
            if blanks == 0 and score >= 0.82:
                continue
            candidates.append((blanks, -score, r, row))

        target = None
        if candidates:
            candidates.sort(reverse=True)
            target = candidates[0][2]

        if target:
            write_missing_row(ws, target, ref, preferred_drivers, d_value)
            used_update_targets.add(target)
            updated_rows.append((target, year, ref["wiki_driver"], ref["wiki_entrant"]))
        else:
            skipped_rows.append((ref, d_value))

    for ref, d_value in skipped_rows:
        year = int(ref["year"])
        insert_at = find_insert_row(ws, year, ref["driver_key"], d_value)
        first_year_row = min(year_rows(ws, year) or [2])
        template_row = insert_at - 1 if insert_at > first_year_row else insert_at + 1
        ws.insert_rows(insert_at)
        copy_row_format(ws, template_row, insert_at)
        write_missing_row(ws, insert_at, ref, preferred_drivers, d_value)
        inserted_rows.append((insert_at, year, ref["wiki_driver"], ref["wiki_entrant"]))

    pc_blank_to_zero = 0
    links_removed = 0
    for r in range(2, ws.max_row + 1):
        year = ws.cell(r, COL_YEAR).value
        if not isinstance(year, int) or year < START_YEAR:
            continue
        pc = ws.cell(r, COL_PC).value
        brand = clean(ws.cell(r, 7).value)
        has_any_link = any(ws.cell(r, c).hyperlink for c in range(1, MAX_COL + 1))
        if pc in (None, "") and not brand:
            ws.cell(r, COL_PC).value = 0
            pc_blank_to_zero += 1
            pc = 0
        if pc in (0, "0"):
            if has_any_link:
                links_removed += 1
            clear_row_links(ws, r)

    if ws.auto_filter and ws.auto_filter.ref:
        ws.auto_filter.ref = f"A1:P{ws.max_row}"

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    # A second pass after save clears hyperlink relationships that Excel/OpenXML
    # can keep around after row insertion shifts existing linked cells.
    wb = load_workbook(OUTPUT)
    ws = wb["Overview"]
    for r in range(2, ws.max_row + 1):
        year = ws.cell(r, COL_YEAR).value
        if isinstance(year, int) and year >= START_YEAR and ws.cell(r, COL_PC).value in (0, "0"):
            clear_row_links(ws, r)
    wb.save(OUTPUT)

    print(f"Saved: {OUTPUT.resolve()}")
    print(f"Inserted rows: {len(inserted_rows)}")
    print(f"Updated existing rows: {len(updated_rows)}")
    print(f"Pc blank -> 0 rows: {pc_blank_to_zero}")
    print(f"Pc=0 rows with links removed: {links_removed}")


if __name__ == "__main__":
    main()
