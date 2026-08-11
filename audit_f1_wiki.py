from __future__ import annotations

import csv
import re
import time
import unicodedata
import urllib.request
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from urllib.error import HTTPError

from lxml import html
from openpyxl import load_workbook


WORKBOOK = Path("originalni_soubor") / "Diecast 2026.xlsx"
OUT_DIR = Path("outputs") / "wiki_audit"
START_YEAR = 1976
END_YEAR = 2025


def text_of(node) -> str:
    return " ".join(" ".join(node.itertext()).split())


def clean_wiki_text(value: str | None) -> str:
    if not value:
        return ""
    value = re.sub(r"\[\s*[a-zA-Z0-9]+\s*\]", "", value)
    value = value.replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value).strip()
    return value


def norm(value: object) -> str:
    if value is None:
        return ""
    value = str(value).replace("\xa0", " ")
    value = re.sub(r"([a-z])([A-Z])", r"\1 \2", value)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower()
    value = value.replace("&", " and ")
    value = value.replace("–", "-").replace("—", "-")
    value = re.sub(r"\[[^\]]+\]", " ", value)
    value = re.sub(r"\b(f1|formula one|formula 1|racing|team|scuderia)\b", " ", value)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def driver_key(value: object) -> str:
    n = norm(value)
    n = re.sub(r"\b(jr|junior|sr|ii|iii)\b", " ", n)
    tokens = [t for t in re.split(r"\s+", n) if t]
    return " ".join(sorted(tokens))


def ratio(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.92
    return SequenceMatcher(None, a, b).ratio()


def driver_similarity(a: object, b: object) -> float:
    return max(ratio(driver_key(a), driver_key(b)), ratio(norm(a), norm(b)))


def fetch_doc(year: int):
    urls = [
        f"https://en.wikipedia.org/wiki/{year}_Formula_One_World_Championship",
        f"https://en.wikipedia.org/wiki/{year}_Formula_One_season",
    ]
    last_error = None
    for url in urls:
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": "Mozilla/5.0 (Codex local F1 spreadsheet audit)"}
            )
            data = urllib.request.urlopen(req, timeout=45).read()
            return url, html.fromstring(data)
        except HTTPError as exc:
            last_error = exc
            if exc.code != 404:
                break
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            break
    raise RuntimeError(f"Cannot fetch {year}: {last_error}")


def expand_table(table):
    rows = table.xpath(".//tr")
    active: dict[int, tuple[int, object]] = {}
    matrix = []
    for tr in rows:
        row = []
        col = 0
        for c, (remaining, cell) in sorted(list(active.items())):
            while len(row) <= c:
                row.append(None)
            row[c] = cell
            if remaining <= 1:
                del active[c]
            else:
                active[c] = (remaining - 1, cell)
        for cell in tr.xpath("./th|./td"):
            while col < len(row) and row[col] is not None:
                col += 1
            rowspan = int(cell.get("rowspan", "1") or "1")
            colspan = int(cell.get("colspan", "1") or "1")
            for offset in range(colspan):
                idx = col + offset
                while len(row) <= idx:
                    row.append(None)
                row[idx] = cell
                if rowspan > 1:
                    active[idx] = (rowspan - 1, cell)
            col += colspan
        matrix.append(row)
    width = max((len(r) for r in matrix), default=0)
    for r in matrix:
        r.extend([None] * (width - len(r)))
    return matrix


def combined_headers(matrix, header_rows: int):
    headers = []
    width = max((len(r) for r in matrix[:header_rows]), default=0)
    for col in range(width):
        parts = []
        seen = set()
        for row in matrix[:header_rows]:
            cell = row[col] if col < len(row) else None
            if cell is None:
                continue
            t = clean_wiki_text(text_of(cell))
            if t and t not in seen:
                parts.append(t)
                seen.add(t)
        headers.append(" / ".join(parts))
    return headers


def find_entries_table(doc):
    candidates = []
    for table in doc.xpath('//table[contains(@class,"wikitable")]'):
        t = text_of(table)
        score = 0
        if "Entrant" in t:
            score += 5
        if "Constructor" in t:
            score += 3
        if "Driver" in t:
            score += 3
        if "Practice drivers" in t:
            score -= 8
        if "Race drivers" in t:
            score += 2
        if score >= 8:
            candidates.append((score, table))
    if not candidates:
        return None
    return sorted(candidates, key=lambda x: x[0], reverse=True)[0][1]


def find_standings_table(doc):
    best = None
    best_score = 0
    for table in doc.xpath('//table[contains(@class,"wikitable")]'):
        t = text_of(table)
        if "Driver" not in t or ("Points" not in t and "Pts" not in t):
            continue
        score = 0
        if "Pos" in t:
            score += 3
        score += min(t.count("Grand Prix"), 2)
        score += len(table.xpath(".//tr"))
        if score > best_score:
            best = table
            best_score = score
    return best


def driver_names_from_cell(cell):
    if cell is None:
        return []
    names = []
    for a in cell.xpath(".//a[@href]"):
        name = clean_wiki_text(text_of(a))
        href = a.get("href") or ""
        if not name or name.startswith("["):
            continue
        if "/wiki/" not in href:
            continue
        if name in {"Driver", "Driver name", "No.", "Constructor"}:
            continue
        if any(bad in name.lower() for bad in ("citation needed", "failed verification", "source", "sources")):
            continue
        if len(name) <= 2:
            continue
        names.append(name)
    out = []
    seen = set()
    for name in names:
        key = driver_key(name)
        if key not in seen:
            out.append(name)
            seen.add(key)
    return out


def parse_entries(year: int, doc):
    table = find_entries_table(doc)
    if table is None:
        return []
    matrix = expand_table(table)
    header_start = 0
    for i, row in enumerate(matrix):
        row_text = " ".join(clean_wiki_text(text_of(c)) for c in row if c is not None)
        if "Driver" in row_text and ("Entrant" in row_text or "Constructor" in row_text):
            header_start = i
            break
    matrix = matrix[header_start:]
    header_rows = 2 if any("Race drivers" in clean_wiki_text(text_of(c)) for c in matrix[0] if c is not None) else 1
    headers = combined_headers(matrix, header_rows)
    lowered = [h.lower() for h in headers]

    def find_col(*needles):
        for i, h in enumerate(lowered):
            if all(n in h for n in needles):
                return i
        return None

    entrant_col = find_col("entrant")
    constructor_col = find_col("constructor")
    chassis_col = find_col("chassis")
    driver_col = find_col("driver")
    no_col = next((i for i, h in enumerate(lowered) if re.search(r"(^|/|\s)no\.?($|/|\s)", h)), None)
    rounds_col = find_col("rounds")
    if driver_col is None:
        return []

    records = []
    for row in matrix[header_rows:]:
        if not row or all(c is None or not clean_wiki_text(text_of(c)) for c in row):
            continue
        driver_cell = row[driver_col] if driver_col < len(row) else None
        drivers = driver_names_from_cell(driver_cell)
        if not drivers:
            continue
        no_text = clean_wiki_text(text_of(row[no_col])) if no_col is not None and no_col < len(row) and row[no_col] is not None else ""
        no_values = re.findall(r"\d+", no_text)
        rounds = clean_wiki_text(text_of(row[rounds_col])) if rounds_col is not None and rounds_col < len(row) and row[rounds_col] is not None else ""
        entrant = clean_wiki_text(text_of(row[entrant_col])) if entrant_col is not None and entrant_col < len(row) else ""
        constructor = clean_wiki_text(text_of(row[constructor_col])) if constructor_col is not None and constructor_col < len(row) else ""
        chassis = clean_wiki_text(text_of(row[chassis_col])) if chassis_col is not None and chassis_col < len(row) else ""
        for idx, driver in enumerate(drivers):
            number = no_values[idx] if len(no_values) == len(drivers) else no_text
            records.append(
                {
                    "year": year,
                    "wiki_driver": driver,
                    "driver_key": driver_key(driver),
                    "wiki_entrant": entrant,
                    "wiki_constructor": constructor,
                    "wiki_chassis": chassis,
                    "wiki_no": number,
                    "wiki_rounds": rounds,
                }
            )
    dedup = {}
    for rec in records:
        key = (rec["driver_key"], norm(rec["wiki_entrant"]), norm(rec["wiki_constructor"]))
        dedup.setdefault(key, rec)
    return list(dedup.values())


def parse_driver_order(year: int, doc):
    table = find_standings_table(doc)
    if table is None:
        return {}
    matrix = expand_table(table)
    header_start = 0
    for i, row in enumerate(matrix):
        row_text = " ".join(clean_wiki_text(text_of(c)) for c in row if c is not None)
        if "Driver" in row_text and ("Pos" in row_text or "Points" in row_text or "Pts" in row_text):
            header_start = i
            break
    matrix = matrix[header_start:]
    header_rows = 1
    headers = combined_headers(matrix, header_rows)
    lowered = [h.lower() for h in headers]
    pos_col = next((i for i, h in enumerate(lowered) if h.startswith("pos")), None)
    driver_col = next((i for i, h in enumerate(lowered) if "driver" in h), None)
    if driver_col is None:
        return {}
    order = {}
    for row in matrix[header_rows:]:
        if driver_col >= len(row) or row[driver_col] is None:
            continue
        names = driver_names_from_cell(row[driver_col])
        if not names:
            continue
        pos = clean_wiki_text(text_of(row[pos_col])) if pos_col is not None and pos_col < len(row) else ""
        for name in names:
            order.setdefault(driver_key(name), {"wiki_driver": name, "wiki_pos": pos, "year": year})
    return order


def load_sheet():
    wb = load_workbook(WORKBOOK, data_only=False)
    ws = wb["Overview"]
    rows_by_year = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        year = ws.cell(r, 1).value
        if not isinstance(year, int) or year < START_YEAR or year > END_YEAR:
            continue
        driver = ws.cell(r, 6).value
        if not driver:
            continue
        rec = {
            "row": r,
            "year": year,
            "team": ws.cell(r, 2).value or "",
            "car": ws.cell(r, 3).value or "",
            "type": ws.cell(r, 4).value or "",
            "driver": driver,
            "driver_key": driver_key(driver),
            "D": ws.cell(r, 9).value,
            "Pc": ws.cell(r, 13).value,
            "Code": ws.cell(r, 14).value,
        }
        rows_by_year[year].append(rec)
    return rows_by_year


def best_match_for_ref(ref, sheet_rows):
    candidates = [(driver_similarity(ref["wiki_driver"], r["driver"]), r) for r in sheet_rows]
    candidates = [(ds, r) for ds, r in candidates if ds >= 0.84]
    best = None
    for driver_score, row in candidates:
        team_score = ratio(norm(ref["wiki_entrant"]), norm(row["team"]))
        constructor_score = ratio(norm(ref["wiki_constructor"]), norm(row["car"]))
        chassis_score = ratio(norm(ref["wiki_chassis"]), norm(row["type"]))
        score = max(team_score, constructor_score, chassis_score)
        by = "team" if score == team_score else "constructor" if score == constructor_score else "chassis"
        item = (score, by, team_score, constructor_score, chassis_score, row, driver_score)
        if best is None or item[0] > best[0]:
            best = item
    return best


def best_match_for_sheet(row, refs):
    candidates = [(driver_similarity(row["driver"], r["wiki_driver"]), r) for r in refs]
    candidates = [(ds, r) for ds, r in candidates if ds >= 0.84]
    best = None
    for driver_score, ref in candidates:
        team_score = ratio(norm(ref["wiki_entrant"]), norm(row["team"]))
        constructor_score = ratio(norm(ref["wiki_constructor"]), norm(row["car"]))
        chassis_score = ratio(norm(ref["wiki_chassis"]), norm(row["type"]))
        score = max(team_score, constructor_score, chassis_score)
        by = "team" if score == team_score else "constructor" if score == constructor_score else "chassis"
        item = (score, by, team_score, constructor_score, chassis_score, ref, driver_score)
        if best is None or item[0] > best[0]:
            best = item
    return best


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    sheet = load_sheet()
    ref_by_year = {}
    order_by_year = {}
    sources = []
    for year in range(START_YEAR, END_YEAR + 1):
        url, doc = fetch_doc(year)
        refs = parse_entries(year, doc)
        order = parse_driver_order(year, doc)
        ref_by_year[year] = refs
        order_by_year[year] = order
        sources.append({"year": year, "url": url, "ref_combos": len(refs), "driver_order": len(order)})
        time.sleep(0.15)

    missing = []
    extra = []
    review = []
    order_issues = []
    yearly = []
    threshold = 0.82
    review_threshold = 0.92

    for year in range(START_YEAR, END_YEAR + 1):
        sheet_rows = sheet.get(year, [])
        refs = ref_by_year.get(year, [])
        matched_sheet_rows = set()
        for ref in refs:
            match = best_match_for_ref(ref, sheet_rows)
            if not match or match[0] < threshold:
                missing.append({**ref, "issue": "missing_in_sheet", "best_score": round(match[0], 3) if match else 0})
            else:
                score, by, team_score, constructor_score, chassis_score, row, driver_score = match
                matched_sheet_rows.add(row["row"])
                if score < review_threshold:
                    review.append(
                        {
                            "year": year,
                            "issue": "low_confidence_match",
                            "wiki_driver": ref["wiki_driver"],
                            "wiki_entrant": ref["wiki_entrant"],
                            "wiki_constructor": ref["wiki_constructor"],
                            "wiki_chassis": ref["wiki_chassis"],
                            "sheet_row": row["row"],
                            "sheet_driver": row["driver"],
                            "sheet_team": row["team"],
                            "sheet_car": row["car"],
                            "sheet_type": row["type"],
                            "best_score": round(score, 3),
                            "match_by": by,
                            "team_score": round(team_score, 3),
                            "constructor_score": round(constructor_score, 3),
                            "chassis_score": round(chassis_score, 3),
                            "driver_score": round(driver_score, 3),
                        }
                    )

        sheet_unique = {}
        for row in sheet_rows:
            key = (row["driver_key"], norm(row["team"]), norm(row["car"]), norm(row["type"]))
            sheet_unique.setdefault(key, row)
        for row in sheet_unique.values():
            match = best_match_for_sheet(row, refs)
            if not match or match[0] < threshold:
                extra.append(
                    {
                        "year": year,
                        "issue": "not_found_on_wikipedia_entries",
                        "sheet_row": row["row"],
                        "sheet_driver": row["driver"],
                        "sheet_team": row["team"],
                        "sheet_car": row["car"],
                        "sheet_type": row["type"],
                        "D": row["D"],
                        "Pc": row["Pc"],
                        "Code": row["Code"],
                        "best_score": round(match[0], 3) if match else 0,
                        "best_wiki_driver": match[5]["wiki_driver"] if match else "",
                        "best_wiki_entrant": match[5]["wiki_entrant"] if match else "",
                        "best_wiki_constructor": match[5]["wiki_constructor"] if match else "",
                        "driver_score": round(match[6], 3) if match else 0,
                    }
                )

        order = order_by_year.get(year, {})
        first_by_driver = {}
        for row in sheet_rows:
            first_by_driver.setdefault(row["driver_key"], row)
        for dkey, row in first_by_driver.items():
            if dkey not in order:
                continue
            wiki_pos = order[dkey]["wiki_pos"]
            sheet_d = str(row["D"]).strip() if row["D"] is not None else ""
            if wiki_pos and sheet_d and wiki_pos != sheet_d and wiki_pos.isdigit() and sheet_d.isdigit():
                order_issues.append(
                    {
                        "year": year,
                        "sheet_row": row["row"],
                        "driver": row["driver"],
                        "sheet_D": sheet_d,
                        "wiki_pos": wiki_pos,
                    }
                )

        yearly.append(
            {
                "year": year,
                "wiki_combos": len(refs),
                "sheet_rows": len(sheet_rows),
                "sheet_unique_driver_team_car_type": len(sheet_unique),
                "missing_count": sum(1 for x in missing if x["year"] == year),
                "extra_count": sum(1 for x in extra if x["year"] == year),
                "review_count": sum(1 for x in review if x["year"] == year),
                "order_issue_count": sum(1 for x in order_issues if x["year"] == year),
            }
        )

    def write_csv(name, rows):
        path = OUT_DIR / name
        fieldnames = sorted({k for row in rows for k in row.keys()})
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        return path

    write_csv("sources.csv", sources)
    write_csv("yearly_summary.csv", yearly)
    write_csv("missing_in_sheet.csv", missing)
    write_csv("extra_in_sheet.csv", extra)
    write_csv("review_low_confidence.csv", review)
    write_csv("order_issues.csv", order_issues)
    print(f"OUT_DIR={OUT_DIR.resolve()}")
    print(f"SOURCES={len(sources)} MISSING={len(missing)} EXTRA={len(extra)} REVIEW={len(review)} ORDER={len(order_issues)}")
    print("YEARS_WITH_MISSING", [y["year"] for y in yearly if y["missing_count"]])
    print("YEARS_WITH_EXTRA", [y["year"] for y in yearly if y["extra_count"]])


if __name__ == "__main__":
    main()
